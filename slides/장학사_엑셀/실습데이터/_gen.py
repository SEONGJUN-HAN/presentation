# -*- coding: utf-8 -*-
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
cfg = json.load(open(os.path.join(BASE, "_data.json"), encoding="utf-8"))
L, SEM, OUT = cfg["L"], cfg["sem"], cfg["out"]
FONT = "맑은 고딕"

thin = Side(style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFILL = PatternFill("solid", fgColor="D9D9D9")
F_BASE  = Font(name=FONT, size=10)
F_TITLE = Font(name=FONT, size=13, bold=True)
F_MEMO  = Font(name=FONT, size=9, italic=True, color="666666")
F_HEAD  = Font(name=FONT, size=10, bold=True)
F_BOLD  = Font(name=FONT, size=10, bold=True)
F_SUB   = Font(name=FONT, size=10, italic=True)
CENTER  = Alignment(horizontal="center", vertical="center")
VCENTER = Alignment(vertical="center")
NUMFMT  = "#,##0"


def put(ws, r, c, v, font=F_BASE, align=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    return cell


def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


def head(ws, r, c):
    cell = ws.cell(row=r, column=c)
    cell.font = F_HEAD
    cell.fill = HFILL
    cell.alignment = CENTER


def merge_area_runs(ws, col, r1, r2):
    start = r1
    prev = ws.cell(row=r1, column=col).value
    for r in range(r1 + 1, r2 + 2):
        cur = ws.cell(row=r, column=col).value if r <= r2 else "\x00END"
        if cur != prev:
            if r - 1 > start:
                ws.merge_cells(start_row=start, start_column=col, end_row=r - 1, end_column=col)
                ws.cell(row=start, column=col).alignment = VCENTER
            start = r
            prev = cur


# ---------- Example 1 source: single header row ----------
def build_ex1_source(sc, path):
    wb = Workbook(); ws = wb.active; ws.title = L["sheet"]
    put(ws, 1, 1, sc["title"], F_TITLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    put(ws, 2, 1, sc["memo"], F_MEMO)
    hr = 4
    for c, t in enumerate([L["area"], L["course"], SEM[0], SEM[1]], start=1):
        put(ws, hr, c, t); head(ws, hr, c)
    r = hr + 1; start = r; tot = [0, 0]
    for row in sc["rows"]:
        put(ws, r, 1, row["area"]); put(ws, r, 2, row["course"])
        for s in range(2):
            v = row["v"][s]
            if v is None:
                put(ws, r, 3 + s, L["dash"], align=CENTER)
            else:
                put(ws, r, 3 + s, int(v), fmt=NUMFMT); tot[s] += int(v)
        r += 1
    end = r - 1
    put(ws, r, 1, L["total"], F_BOLD)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    put(ws, r, 3, tot[0], F_BOLD, fmt=NUMFMT); put(ws, r, 4, tot[1], F_BOLD, fmt=NUMFMT)
    merge_area_runs(ws, 1, start, end)
    box(ws, hr, 1, r, 4)
    for c, w in zip("ABCD", [10, 16, 10, 10]):
        ws.column_dimensions[c].width = w
    wb.save(path)


# ---------- Example 2 source: two header rows + subtotals ----------
def build_ex2_source(sc, path):
    wb = Workbook(); ws = wb.active; ws.title = L["sheet"]
    put(ws, 1, 1, sc["title"], F_TITLE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    put(ws, 2, 1, sc["memo"], F_MEMO)
    hr = 4
    put(ws, hr, 1, L["area"]); ws.merge_cells(start_row=hr, start_column=1, end_row=hr+1, end_column=1)
    put(ws, hr, 2, L["course"]); ws.merge_cells(start_row=hr, start_column=2, end_row=hr+1, end_column=2)
    put(ws, hr, 3, SEM[0]); ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=4)
    put(ws, hr, 5, SEM[1]); ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)
    put(ws, hr+1, 3, L["count"]); put(ws, hr+1, 4, L["people"])
    put(ws, hr+1, 5, L["count"]); put(ws, hr+1, 6, L["people"])
    for c in range(1, 7):
        head(ws, hr, c); head(ws, hr + 1, c)
    r = hr + 2; start = r; grand = [0, 0, 0, 0]; rows = sc["rows"]; i = 0
    while i < len(rows):
        area = rows[i]["area"]; a_start = r; sub = [0, 0, 0, 0]
        while i < len(rows) and rows[i]["area"] == area:
            row = rows[i]
            put(ws, r, 1, area); put(ws, r, 2, row["course"])
            vals = [row["c"][0], row["n"][0], row["c"][1], row["n"][1]]
            for k in range(4):
                v = vals[k]
                if v is None:
                    put(ws, r, 3 + k, L["dash"], align=CENTER)
                else:
                    put(ws, r, 3 + k, int(v), fmt=NUMFMT); sub[k] += int(v); grand[k] += int(v)
            r += 1; i += 1
        a_end = r - 1
        if a_end > a_start:
            ws.merge_cells(start_row=a_start, start_column=1, end_row=a_end, end_column=1)
            ws.cell(row=a_start, column=1).alignment = VCENTER
        put(ws, r, 2, L["subtotal"], F_SUB)
        for k in range(4):
            put(ws, r, 3 + k, sub[k], F_SUB, fmt=NUMFMT)
        r += 1
    put(ws, r, 1, L["total"], F_BOLD)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    for k in range(4):
        put(ws, r, 3 + k, grand[k], F_BOLD, fmt=NUMFMT)
    box(ws, hr, 1, r, 6)
    for c, w in zip("ABCDEF", [10, 16, 10, 10, 10, 10]):
        ws.column_dimensions[c].width = w
    wb.save(path)


def write_notes(ws, lines):
    for i, ln in enumerate(lines, start=1):
        f = F_TITLE if i == 1 else F_BASE
        put(ws, i, 1, ln, f)
    ws.column_dimensions["A"].width = 105


# ---------- Answer key Example 1 ----------
def build_ex1_answer(schools, path, notes):
    wb = Workbook(); ds = wb.active; ds.title = L["datasheet"]
    hdr = [L["school"], L["area"], L["course"], L["semester"], L["people"]]
    for c, t in enumerate(hdr, start=1):
        put(ds, 1, c, t); head(ds, 1, c)
    r = 2
    for sc in schools:
        for row in sc["rows"]:
            for s in range(2):
                put(ds, r, 1, sc["short"]); put(ds, r, 2, row["area"])
                put(ds, r, 3, row["course"]); put(ds, r, 4, SEM[s], align=CENTER)
                v = row["v"][s]
                if v is not None: put(ds, r, 5, int(v), fmt=NUMFMT)
                r += 1
    box(ds, 1, 1, r - 1, 5)
    for c, w in zip("ABCDE", [10, 10, 16, 10, 12]):
        ds.column_dimensions[c].width = w

    pv = wb.create_sheet(L["pivotsheet"])
    for c, t in enumerate([L["school"], SEM[0], SEM[1], L["total"]], start=1):
        put(pv, 1, c, t); head(pv, 1, c)
    pr = 2; coltot = [0, 0]
    for sc in schools:
        put(pv, pr, 1, sc["short"])
        rowsum = 0
        for s in range(2):
            val = sum(int(row["v"][s]) for row in sc["rows"] if row["v"][s] is not None)
            put(pv, pr, 2 + s, val, fmt=NUMFMT); rowsum += val; coltot[s] += val
        put(pv, pr, 4, rowsum, F_BOLD, fmt=NUMFMT)
        pr += 1
    put(pv, pr, 1, L["grandtotal"], F_BOLD)
    put(pv, pr, 2, coltot[0], F_BOLD, fmt=NUMFMT); put(pv, pr, 3, coltot[1], F_BOLD, fmt=NUMFMT)
    put(pv, pr, 4, coltot[0] + coltot[1], F_BOLD, fmt=NUMFMT)
    box(pv, 1, 1, pr, 4)
    for c, w in zip("ABCD", [12, 12, 12, 12]):
        pv.column_dimensions[c].width = w

    write_notes(wb.create_sheet(L["notesheet"]), notes)
    wb.active = 0
    wb.save(path)


# ---------- Answer key Example 2 ----------
def build_ex2_answer(schools, path, notes):
    wb = Workbook(); ds = wb.active; ds.title = L["datasheet"]
    hdr = [L["school"], L["area"], L["course"], L["semester"], L["count"], L["people"]]
    for c, t in enumerate(hdr, start=1):
        put(ds, 1, c, t); head(ds, 1, c)
    r = 2
    for sc in schools:
        for row in sc["rows"]:
            for s in range(2):
                put(ds, r, 1, sc["short"]); put(ds, r, 2, row["area"])
                put(ds, r, 3, row["course"]); put(ds, r, 4, SEM[s], align=CENTER)
                cv, nv = row["c"][s], row["n"][s]
                if cv is not None: put(ds, r, 5, int(cv), fmt=NUMFMT)
                if nv is not None: put(ds, r, 6, int(nv), fmt=NUMFMT)
                r += 1
    box(ds, 1, 1, r - 1, 6)
    for c, w in zip("ABCDEF", [10, 10, 16, 10, 10, 12]):
        ds.column_dimensions[c].width = w

    pv = wb.create_sheet(L["pivotsheet"])
    hdr2 = [L["school"], f"{SEM[0]} {L['count']}", f"{SEM[0]} {L['people']}",
            f"{SEM[1]} {L['count']}", f"{SEM[1]} {L['people']}"]
    for c, t in enumerate(hdr2, start=1):
        put(pv, 1, c, t); head(pv, 1, c)
    pr = 2; coltot = [0, 0, 0, 0]
    for sc in schools:
        put(pv, pr, 1, sc["short"])
        vals = []
        for s in range(2):
            cc = sum(int(row["c"][s]) for row in sc["rows"] if row["c"][s] is not None)
            nn = sum(int(row["n"][s]) for row in sc["rows"] if row["n"][s] is not None)
            vals += [cc, nn]
        for k in range(4):
            put(pv, pr, 2 + k, vals[k], fmt=NUMFMT); coltot[k] += vals[k]
        pr += 1
    put(pv, pr, 1, L["grandtotal"], F_BOLD)
    for k in range(4):
        put(pv, pr, 2 + k, coltot[k], F_BOLD, fmt=NUMFMT)
    box(pv, 1, 1, pr, 5)
    for c, w in zip("ABCDE", [10, 14, 14, 14, 14]):
        pv.column_dimensions[c].width = w

    write_notes(wb.create_sheet(L["notesheet"]), notes)
    wb.active = 0
    wb.save(path)


ex1dir = os.path.join(BASE, OUT["ex1dir"]); ex1src = os.path.join(ex1dir, OUT["srcdir"])
ex2dir = os.path.join(BASE, OUT["ex2dir"]); ex2src = os.path.join(ex2dir, OUT["srcdir"])
os.makedirs(ex1src, exist_ok=True); os.makedirs(ex2src, exist_ok=True)

for sc in cfg["ex1"]["schools"]:
    build_ex1_source(sc, os.path.join(ex1src, sc["file"] + ".xlsx"))
for sc in cfg["ex2"]["schools"]:
    build_ex2_source(sc, os.path.join(ex2src, sc["file"] + ".xlsx"))
build_ex1_answer(cfg["ex1"]["schools"], os.path.join(ex1dir, OUT["answer"]), cfg["notes1"])
build_ex2_answer(cfg["ex2"]["schools"], os.path.join(ex2dir, OUT["answer"]), cfg["notes2"])
print("DONE")
